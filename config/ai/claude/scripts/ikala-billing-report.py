#!/usr/bin/env python3
"""Monthly GCP billing report builder for the iKala-resold accounts
(dunqian2 / dunqian5 - ikalatv).

Scans a directory for GCP console "Cost Table" CSV exports, builds one xlsx
workbook per (billing account, month) in the layout of the reference report
202504_敦謙國際智能股份有限公司.xlsx (總表 + per-project sheets + 小計), plus a
cross-month summary workbook. Optionally uploads everything to Google Drive
as native Google Sheets (updates in place on re-run, so links stay stable).

Usage: see ikala-billing-report.sh
"""
import argparse
import csv
import glob
import json
import os
import re
import subprocess
import sys
import urllib.error
import urllib.parse
import urllib.request
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

DEFAULT_SRC = os.path.expanduser("~/Downloads")
DEFAULT_DRIVE_FOLDER = "1Qtk7ZoaDXg4ehWxi8iiUEpFqmd9I2J9U"  # Drive「帳單」資料夾

CSV_SIGNATURE = ["Billing account name", "Billing account ID", "Project ID",
                 "Service description", "SKU description", "SKU ID",
                 "Cost type", "Usage start date", "Usage end date",
                 "Usage amount", "Usage unit", "Unrounded Cost ($)", "Cost ($)"]

CREDIT_ZH = {
    "SPENDING_BASED_DISCOUNT": "折扣(消費折扣)",
    "SUSTAINED_USAGE_DISCOUNT": "折扣(持續使用)",
    "PROMOTION": "折扣(促銷抵免)",
    "FEE_UTILIZATION_OFFSET": "折扣(費用抵銷)",
}

HEADERS = ["帳單帳務 ID", "專案ID", "服務 ID", "服務說明", "SKU ID", "SKU 說明",
           "費用類型", "開始日期", "結束日期", "用量", "用量單位", "合約價"]

HDR_FILL = PatternFill("solid", fgColor="D9E1F2")
SUM_FILL = PatternFill("solid", fgColor="FCE4D6")
RATE_FILL = PatternFill("solid", fgColor="FFF2CC")
BOLD = Font(bold=True)


def fnum(s):
    s = (s or "").replace(",", "").strip()
    return float(s) if s else None


def discover(src):
    """Find cost-table CSVs and group usage rows by (account name, YYYYMM)."""
    groups = {}  # (acct, month) -> {"rows": [...], "total": float|None, "files": [...]}
    for path in sorted(glob.glob(os.path.join(src, "*.csv"))):
        try:
            with open(path, encoding="utf-8-sig", newline="") as fh:
                reader = csv.DictReader(fh)
                if not reader.fieldnames or not all(c in reader.fieldnames for c in CSV_SIGNATURE):
                    continue
                rows = list(reader)
        except (OSError, UnicodeDecodeError):
            continue
        usage = [r for r in rows if r["Cost type"] == "Usage" and r["Billing account name"]]
        if not usage:
            continue
        months = {}
        for r in usage:
            months[r["Usage end date"][:7]] = months.get(r["Usage end date"][:7], 0) + 1
        month = max(months, key=months.get).replace("-", "")
        acct = usage[0]["Billing account name"]
        total = None
        for r in rows:
            if r["Cost type"] == "Total" and not r["Billing account name"]:
                total = fnum(r["Unrounded Cost ($)"]) or fnum(r["Cost ($)"])
        key = (acct, month)
        if key in groups:
            sys.exit(f"error: duplicate export for {acct} {month}:\n"
                     f"  {groups[key]['files'][0]}\n  {path}\n"
                     "remove one of the files and re-run.")
        groups[key] = {"rows": usage, "total": total, "files": [path]}
    return groups


def aggregate(rows):
    """One line per (project, service, sku, credit type): min/max dates, summed usage+cost."""
    agg = {}
    for r in rows:
        key = (r["Project ID"], r["Service ID"], r["SKU ID"], r["Credit type"])
        a = agg.setdefault(key, {
            "acct": r["Billing account ID"],
            "proj_id": r["Project ID"], "proj_name": r["Project name"],
            "svc_id": r["Service ID"], "svc": r["Service description"],
            "sku_id": r["SKU ID"], "sku": r["SKU description"],
            "credit": r["Credit type"],
            "start": r["Usage start date"], "end": r["Usage end date"],
            "usage": 0.0, "has_usage": False, "unit": r["Usage unit"],
            "cost": 0.0,
        })
        if r["Usage start date"] and (not a["start"] or r["Usage start date"] < a["start"]):
            a["start"] = r["Usage start date"]
        if r["Usage end date"] and r["Usage end date"] > a["end"]:
            a["end"] = r["Usage end date"]
        u = fnum(r["Usage amount"])
        if u is not None:
            a["usage"] += u
            a["has_usage"] = True
        if r["Usage unit"]:
            a["unit"] = r["Usage unit"]
        a["cost"] += fnum(r["Unrounded Cost ($)"]) or 0.0
    out = list(agg.values())
    out.sort(key=lambda a: (a["proj_name"] or "~", a["svc"], a["sku"], a["credit"]))
    return out


def sheet_title(name):
    if not name or name.startswith("[Charges"):
        return "Other"
    t = re.sub(r"[\[\]:*?/\\]", "", name)
    parts = re.split(r"[-_ ]+", t)
    t = "".join(p[:1].upper() + p[1:] for p in parts if p)
    return t[:31]


def to_date(s):
    return datetime.strptime(s, "%Y-%m-%d") if s else None


def write_detail_sheet(ws, lines, rate):
    for c, h in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = BOLD
        cell.fill = HDR_FILL
    last = len(lines) + 1
    for i, a in enumerate(lines, 2):
        ftype = CREDIT_ZH.get(a["credit"], a["credit"]) if a["credit"] else "用量"
        ws.cell(row=i, column=1, value=a["acct"])
        ws.cell(row=i, column=2, value=a["proj_id"] or "(無專案)")
        ws.cell(row=i, column=3, value=a["svc_id"])
        ws.cell(row=i, column=4, value=a["svc"])
        ws.cell(row=i, column=5, value=a["sku_id"])
        ws.cell(row=i, column=6, value=a["sku"])
        ws.cell(row=i, column=7, value=ftype)
        for col, d in ((8, a["start"]), (9, a["end"])):
            c = ws.cell(row=i, column=col, value=to_date(d))
            c.number_format = "yyyy-mm-dd"
        if a["has_usage"]:
            c = ws.cell(row=i, column=10, value=round(a["usage"], 3))
            c.number_format = "#,##0.###"
            ws.cell(row=i, column=11, value=a["unit"])
        c = ws.cell(row=i, column=12, value=round(a["cost"], 2))
        c.number_format = "#,##0.00"
    labels = ["用量金額(USD)", "調整金額(USD)", "匯率", "帳款金額合計 (NTD)"]
    for r, lab in enumerate(labels, 1):
        ws.cell(row=r, column=13, value=lab).font = BOLD
        ws.cell(row=r, column=13).fill = SUM_FILL
    n1 = ws.cell(row=1, column=14, value=f"=ROUND(SUM(L2:L{max(last, 2)}),2)")
    n1.number_format = "#,##0.00"
    n1.font = BOLD
    ws.cell(row=2, column=14, value=0).number_format = "#,##0.00"
    # GOOGLEFINANCE only works after conversion to Google Sheets; --rate pins a fixed invoice rate
    rc = ws.cell(row=3, column=14, value=rate if rate else '=GOOGLEFINANCE("CURRENCY:USDTWD")')
    rc.number_format = "#,##0.00"
    rc.fill = RATE_FILL
    n4 = ws.cell(row=4, column=14, value="=ROUND(N1*N3,2)")
    n4.number_format = "#,##0.00"
    n4.font = BOLD
    widths = [22, 22, 16, 26, 16, 58, 16, 12, 12, 14, 18, 12, 20, 14]
    for c, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:L{max(last, 2)}"


def write_subtotal_sheet(ws, lines, month, acct_name, csv_total):
    ws.cell(row=1, column=1, value=f"{acct_name}  {month[:4]}-{month[4:]} 小計").font = Font(bold=True, size=13)
    by_proj, by_svc = {}, {}
    credits = gross = 0.0
    for a in lines:
        p = a["proj_name"] or "(無專案)"
        by_proj[p] = by_proj.get(p, 0.0) + a["cost"]
        by_svc[a["svc"]] = by_svc.get(a["svc"], 0.0) + a["cost"]
        if a["credit"]:
            credits += a["cost"]
        else:
            gross += a["cost"]
    total = gross + credits
    r = 3
    for block, data in (("按專案小計", by_proj), ("按服務小計", by_svc)):
        ws.cell(row=r, column=1, value=block).font = BOLD
        r += 1
        for c, h in enumerate([block[2:4], "金額(USD)", "占比"], 1):
            ws.cell(row=r, column=c, value=h).font = BOLD
            ws.cell(row=r, column=c).fill = HDR_FILL
        r += 1
        for k, v in sorted(data.items(), key=lambda kv: -kv[1]):
            ws.cell(row=r, column=1, value=k)
            ws.cell(row=r, column=2, value=round(v, 2)).number_format = "#,##0.00"
            ws.cell(row=r, column=3, value=v / total if total else 0).number_format = "0.0%"
            r += 1
        ws.cell(row=r, column=1, value="合計").font = BOLD
        c = ws.cell(row=r, column=2, value=round(total, 2))
        c.number_format = "#,##0.00"
        c.font = BOLD
        r += 2
    for lab, val in [("用量金額小計 (USD)", gross), ("折扣/抵免合計 (USD)", credits),
                     ("本月帳款合計 (USD)", total), ("CSV Total 核對值 (USD)", csv_total)]:
        ws.cell(row=r, column=1, value=lab).font = BOLD
        ws.cell(row=r, column=2, value=round(val, 2) if val is not None else None).number_format = "#,##0.00"
        r += 1
    ws.column_dimensions["A"].width = 46
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 10
    return gross, credits, total


def span_label(months):
    months = sorted(months)
    if len(months) == 1:
        return months[0]
    seq = [int(m[:4]) * 12 + int(m[4:]) for m in months]
    consecutive = seq == list(range(seq[0], seq[0] + len(seq)))
    if consecutive and len(months) == 3 and int(months[0][4:]) % 3 == 1:
        return f"{months[0][:4]}Q{(int(months[0][4:]) - 1) // 3 + 1}"
    return f"{months[0]}-{months[-1]}"


def build(groups, out_dir, rate):
    os.makedirs(out_dir, exist_ok=True)
    summary = []
    for (acct_name, month), g in sorted(groups.items()):
        lines = aggregate(g["rows"])
        wb = Workbook()
        write_detail_sheet(wb.active, lines, rate)
        wb.active.title = "總表"
        proj_cost = {}
        for a in lines:
            proj_cost[a["proj_name"]] = proj_cost.get(a["proj_name"], 0.0) + a["cost"]
        named = sorted([p for p in proj_cost if p and not p.startswith("[Charges")],
                       key=lambda p: -proj_cost[p])
        for p in named:
            write_detail_sheet(wb.create_sheet(sheet_title(p)),
                               [a for a in lines if a["proj_name"] == p], rate)
        other = [a for a in lines if not a["proj_name"] or a["proj_name"].startswith("[Charges")]
        if other:
            write_detail_sheet(wb.create_sheet("Other"), other, rate)
        gross, credits, total = write_subtotal_sheet(
            wb.create_sheet("小計"), lines, month, acct_name, g["total"])
        slug = acct_name.replace(" - ", "-").replace(" ", "")
        fname = f"{month}_{slug}.xlsx"
        wb.save(os.path.join(out_dir, fname))
        diff = abs(total - g["total"]) if g["total"] is not None else None
        if diff is not None and diff > 0.02:
            sys.exit(f"error: {fname} total {total:.2f} does not match "
                     f"CSV Total {g['total']:.2f} — aborting.")
        summary.append((acct_name, month, gross, credits, total, fname))
        print(f"  {fname}: rows={len(lines)} net={total:.2f} "
              f"(csv total check: {'ok' if diff is not None else 'n/a'})")
    label = span_label({m for _, m in groups})
    wb = Workbook()
    ws = wb.active
    ws.title = "彙總"
    ws.cell(row=1, column=1, value=f"ikalatv 帳單彙總 {label}").font = Font(bold=True, size=14)
    for c, h in enumerate(["帳單帳戶", "月份", "用量金額(USD)", "折扣/抵免(USD)",
                           "帳款合計(USD)", "報表檔案"], 1):
        ws.cell(row=3, column=c, value=h).font = BOLD
        ws.cell(row=3, column=c).fill = HDR_FILL
    r = 4
    for acct_name, month, gross, credits, total, fname in summary:
        ws.cell(row=r, column=1, value=acct_name)
        ws.cell(row=r, column=2, value=f"{month[:4]}-{month[4:]}")
        for c, v in ((3, gross), (4, credits), (5, total)):
            ws.cell(row=r, column=c, value=round(v, 2)).number_format = "#,##0.00"
        ws.cell(row=r, column=6, value=fname)
        r += 1
    for acct in sorted({s[0] for s in summary}):
        sub = [s for s in summary if s[0] == acct]
        ws.cell(row=r, column=1, value=f"{acct} 小計").font = BOLD
        for c, idx in ((3, 2), (4, 3), (5, 4)):
            cell = ws.cell(row=r, column=c, value=round(sum(s[idx] for s in sub), 2))
            cell.number_format = "#,##0.00"
            cell.font = BOLD
        r += 1
    ws.cell(row=r, column=1, value="全帳戶總計").font = BOLD
    for c, idx in ((3, 2), (4, 3), (5, 4)):
        cell = ws.cell(row=r, column=c, value=round(sum(s[idx] for s in summary), 2))
        cell.number_format = "#,##0.00"
        cell.font = BOLD
    r += 2
    for n in ["說明:",
              "1. 資料來源: GCP 控制台 Cost Table CSV(SKU 層級, 依 專案+服務+SKU 彙總整月)。",
              "2. 各月檔案內含: 總表 / 各專案分頁 / 小計。折扣以負數列入, SUM 即為帳款淨額。",
              "3. 各分頁 N3 匯率(黃底)預設 =GOOGLEFINANCE(\"CURRENCY:USDTWD\") 即期匯率,"
              " 轉成 Google Sheet 後生效; 要對發票請改填 iKala 當月匯率(或產表時用 --rate)。",
              "4. 每檔小計頁底部的「CSV Total 核對值」為原始 CSV 內建總額, 產表時已核對一致。"]:
        ws.cell(row=r, column=1, value=n)
        r += 1
    for col, w in (("A", 30), ("B", 10), ("C", 16), ("D", 16), ("E", 16), ("F", 30)):
        ws.column_dimensions[col].width = w
    sname = f"00_{label}_彙總.xlsx"
    wb.save(os.path.join(out_dir, sname))
    print(f"  {sname}")
    return label


def drive_api(token, url, data=None, headers=None, method=None):
    req = urllib.request.Request(url, data=data, method=method)
    req.add_header("Authorization", f"Bearer {token}")
    for k, v in (headers or {}).items():
        req.add_header(k, v)
    with urllib.request.urlopen(req) as resp:
        return json.load(resp)


def upload(out_dir, folder_id):
    token = subprocess.run(["gcloud", "auth", "print-access-token"],
                           capture_output=True, text=True, check=True).stdout.strip()
    try:
        folder = drive_api(token, "https://www.googleapis.com/drive/v3/files/"
                           f"{folder_id}?fields=id,name&supportsAllDrives=true")
    except urllib.error.HTTPError as e:
        if e.code == 403:
            sys.exit("error: gcloud token has no Drive scope. Run:\n"
                     "  gcloud auth login --enable-gdrive-access")
        raise
    print(f"uploading to Drive folder「{folder['name']}」...")
    boundary = "xlsx_upload_boundary"
    xlsx_mime = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    for path in sorted(glob.glob(os.path.join(out_dir, "*.xlsx"))):
        name = os.path.splitext(os.path.basename(path))[0]
        q = urllib.parse.quote(f"name='{name}' and '{folder_id}' in parents and trashed=false")
        existing = drive_api(token, "https://www.googleapis.com/drive/v3/files"
                             f"?q={q}&supportsAllDrives=true&includeItemsFromAllDrives=true"
                             "&fields=files(id)")["files"]
        with open(path, "rb") as fh:
            content = fh.read()
        if existing:
            meta = json.dumps({"mimeType": "application/vnd.google-apps.spreadsheet"})
            url = (f"https://www.googleapis.com/upload/drive/v3/files/{existing[0]['id']}"
                   "?uploadType=multipart&supportsAllDrives=true&fields=id,name,webViewLink")
            method = "PATCH"
        else:
            meta = json.dumps({"name": name, "parents": [folder_id],
                               "mimeType": "application/vnd.google-apps.spreadsheet"})
            url = ("https://www.googleapis.com/upload/drive/v3/files"
                   "?uploadType=multipart&supportsAllDrives=true&fields=id,name,webViewLink")
            method = "POST"
        body = (f"--{boundary}\r\nContent-Type: application/json; charset=UTF-8\r\n\r\n{meta}\r\n"
                f"--{boundary}\r\nContent-Type: {xlsx_mime}\r\n\r\n").encode() \
            + content + f"\r\n--{boundary}--".encode()
        r = drive_api(token, url, data=body, method=method,
                      headers={"Content-Type": f"multipart/related; boundary={boundary}"})
        action = "updated" if existing else "created"
        link = r.get("webViewLink", f"https://docs.google.com/spreadsheets/d/{r['id']}")
        print(f"  {action}: {r['name']} {link}")


def main():
    ap = argparse.ArgumentParser(description=__doc__.splitlines()[0])
    ap.add_argument("--src", default=DEFAULT_SRC, help="dir with cost-table CSVs (default: ~/Downloads)")
    ap.add_argument("--out", help="output dir (default: <src>/ikalatv_帳單報表_<span>)")
    ap.add_argument("--months", help="only these months, comma-separated YYYYMM (default: all found)")
    ap.add_argument("--rate", type=float, help="exchange rate to fill into N3 (default: leave blank)")
    ap.add_argument("--upload", nargs="?", const=DEFAULT_DRIVE_FOLDER, metavar="FOLDER_ID",
                    help="upload to Drive as Google Sheets (default folder:「帳單」); updates in place on re-run")
    args = ap.parse_args()

    groups = discover(args.src)
    if args.months:
        wanted = {m.strip().replace("-", "") for m in args.months.split(",")}
        groups = {k: v for k, v in groups.items() if k[1] in wanted}
    if not groups:
        sys.exit(f"error: no cost-table CSVs found in {args.src}")
    for (acct, month), g in sorted(groups.items()):
        print(f"found: {acct} {month} ({len(g['rows'])} rows, {os.path.basename(g['files'][0])})")
    label = span_label({m for _, m in groups})
    out_dir = args.out or os.path.join(args.src, f"ikalatv_帳單報表_{label}")
    print(f"building into {out_dir}")
    build(groups, out_dir, args.rate)
    if args.upload:
        upload(out_dir, args.upload)
    print("done")


if __name__ == "__main__":
    main()
